import os
import pprint
import time

import numpy as np
import torch
from Modules.Loss import (BendingEnergyMetric, DiceCoefficient,
                          JacobianDeterminantMetric, SurfaceDistanceFromSeg)
from openpyxl import Workbook, worksheet
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import csv
import pandas as pd


class MetricTest:
    def __init__(self):
        self.details = {
            'York': {},
            'ACDC': {},
            'MICCAI': {},
            'M&M': {},
        }
        self.info = {
            'York': {
                'LvMyo': [1],
                'LvBp': [2],
                'Lv': [1, 2]
            },
            'ACDC': {
                'Rv': [1],
                'LvMyo': [2],
                'LvBp': [3],
                'Lv': [2, 3],
                # 'LvRv': [1, 2, 3]
            },
            'MICCAI': {
                'LvBp': [1]
            },
            'M&M': {
                'LvBp': [1],
                'LvMyo': [2],
                'Rv': [3],
                'Lv': [1, 2]
            },
        }
        self.dice_estimate = DiceCoefficient()
        self.surface_dist_estimate = SurfaceDistanceFromSeg()
        self.be_estimate = BendingEnergyMetric()
        self.jacobian_estiamte = JacobianDeterminantMetric()

    def getDatasetName(self, case_no):
        if case_no <= 33:
            return 'York'
        elif case_no > 33 and case_no <= 78:
            return 'MICCAI'
        elif case_no > 78 and case_no <= 228:
            return 'ACDC'
        else:
            return 'M&M'

    def testMetrics(self, src, wraped_src, tgt, wraped_tgt, resolution,
                    case_no, slc_idx):
        dataset_name = self.getDatasetName(case_no)
        self.details[dataset_name][case_no] = {}
        self.details[dataset_name][case_no]['slc_idx'] = [
            t.item() for t in slc_idx
        ]

        for key_name in self.info[dataset_name]:
            selected_src = torch.zeros_like(src)
            selected_wraped_src = torch.zeros_like(wraped_src)
            selected_tgt = torch.zeros_like(tgt)
            selected_wraped_tgt = torch.zeros_like(wraped_tgt)
            for v in self.info[dataset_name][key_name]:
                selected_src += (src == v)
                selected_wraped_src += (wraped_src == v)
                selected_tgt += (tgt == v)
                selected_wraped_tgt += (wraped_tgt == v)
            self.details[dataset_name][case_no][key_name] = {
                'undef':
                self.testOnePair(selected_src, selected_tgt, resolution),
                'ed_to_es':
                self.testOnePair(selected_wraped_src, selected_tgt,
                                 resolution),
                'es_to_ed':
                self.testOnePair(selected_wraped_tgt, selected_src, resolution)
            }

    def testOnePair(self, seg_pred_batch, seg_gt_batch, resolution):
        dice_result = self.dice_estimate(seg_pred_batch,
                                         seg_gt_batch).cpu().numpy()
        dice_result = dice_result[np.logical_not(np.isnan(dice_result))]
        # dice_result = dice_result[np.logical_not(dice_result==0)]
        seg_gt_batch = seg_gt_batch.cpu().numpy()[:, 0]
        seg_pred_batch = seg_pred_batch.cpu().numpy()[:, 0]
        apd_result = []
        hd_result = []
        APDgt_pred_result = []
        APDpred_gt_result = []
        for seg_gt, seg_pred in zip(seg_gt_batch, seg_pred_batch):
            # print(seg_gt.shape)
            surface_dist = self.surface_dist_estimate.compute_surface_distances(
                seg_gt, seg_pred, resolution)
            apd = self.surface_dist_estimate.compute_average_surface_distance(
                surface_dist)
            sym_apd = (apd[0] + apd[1]) / 2
            hd = self.surface_dist_estimate.compute_robust_hausdorff(
                surface_dist, 95)
            if not np.isposinf(sym_apd) and not np.isposinf(hd):
                apd_result.append(sym_apd)
                hd_result.append(hd)
                APDgt_pred_result.append(apd[0])
                APDpred_gt_result.append(apd[1])
            else:
                apd_result.append(10)
                hd_result.append(10)
                APDgt_pred_result.append(10)
                APDpred_gt_result.append(10)

        return {
            'Dice': dice_result,
            'SymAPD': apd_result,
            'HD': hd_result,
            'APDgt_pred': APDgt_pred_result,
            'APDpred_gt': APDpred_gt_result
        }

    def testFlow(self, flows_t, flowt_s, case_no):
        dataset_name = self.getDatasetName(case_no)
        bes_t_result = self.be_estimate(flows_t).cpu().numpy()
        jacobians_t_result = torch.sum(self.jacobian_estiamte(flows_t) <= 0,
                                       dim=[1, 2]).cpu().numpy()
        bet_s_result = self.be_estimate(flowt_s).cpu().numpy()
        jacobiant_s_result = torch.sum(self.jacobian_estiamte(flowt_s) <= 0,
                                       dim=[1, 2]).cpu().numpy()
        self.details[dataset_name][case_no]['flow'] = {
            'ed_to_es': {
                'BE': bes_t_result * 1e4,
                'Jacobian': jacobians_t_result
            },
            'es_to_ed': {
                'BE': bet_s_result * 1e4,
                'Jacobian': jacobiant_s_result
            }
        }

    def meanByAnotomicalOfDataset(self):
        # average along with the same anotomical of dataset
        # first average in the same case
        self.mean_by_anotomical = {}

        for dataset_name in self.details:
            dataset = self.details[dataset_name]
            self.mean_by_anotomical[dataset_name] = {}
            for anotomical in self.info[dataset_name]:
                self.mean_by_anotomical[dataset_name][anotomical] = {}
                for direction in ['undef', 'ed_to_es', 'es_to_ed']:
                    self.mean_by_anotomical[dataset_name][anotomical][
                        direction] = {}
                    for metric_name in [
                            'Dice', 'SymAPD', 'HD', 'APDgt_pred', 'APDpred_gt'
                    ]:
                        case_mean = [
                            np.mean(dataset[case_no][anotomical][direction]
                                    [metric_name]) for case_no in dataset
                        ]
                        self.mean_by_anotomical[dataset_name][anotomical][
                            direction][metric_name] = {
                                'mean': np.mean(case_mean),
                                'std': np.std(case_mean)
                            }

    def meanByDataset(self):
        # first average in a case
        # then average along with cases

        self.mean_by_dataset = {}

        for dataset_name in self.details:
            dataset = self.details[dataset_name]
            self.mean_by_dataset[dataset_name] = {}
            for direction in ['undef', 'ed_to_es', 'es_to_ed']:
                self.mean_by_dataset[dataset_name][direction] = {}
                for metric_name in [
                        'Dice', 'SymAPD', 'HD', 'APDgt_pred', 'APDpred_gt'
                ]:
                    case_mean = []
                    for case_no in dataset:
                        case = dataset[case_no]
                        case_item = np.array([
                            case[anotomical][direction][metric_name]
                            for anotomical in self.info[dataset_name]
                        ])
                        case_mean.append(
                            np.mean(np.mean(case_item, axis=0), axis=0))
                    self.mean_by_dataset[dataset_name][direction][
                        metric_name] = {
                            'mean': np.mean(case_mean),
                            'std': np.std(case_mean)
                        }
            # flow
            for direction in ['ed_to_es', 'es_to_ed']:
                for metric_name in ['BE', 'Jacobian']:
                    case_mean = [
                        np.mean(
                            dataset[case_no]['flow'][direction][metric_name])
                        for case_no in dataset
                    ]
                    self.mean_by_dataset[dataset_name][direction][
                        metric_name] = {
                            'mean': np.mean(case_mean),
                            'std': np.std(case_mean)
                        }

    def meanByAll(self):
        # 1. average along with anotomical in a slice
        # 2. average along with slice
        # 3. average along with case_no
        self.mean_by_all = {}

        for direction in ['undef', 'ed_to_es', 'es_to_ed']:
            self.mean_by_all[direction] = {}
            for metric_name in [
                    'Dice', 'SymAPD', 'HD', 'APDgt_pred', 'APDpred_gt'
            ]:
                case_mean = []
                for dataset_name in self.details:
                    dataset = self.details[dataset_name]
                    for case_no in dataset:
                        case = dataset[case_no]
                        case_item = np.array([
                            case[anotomical][direction][metric_name]
                            for anotomical in self.info[dataset_name]
                        ])
                        # 拿到某个数据集的所有case（实际上，所有数据集的case都被保存在相同的list中，也就是最终对所有数据集的所有case求平均），
                        # 对单个case下的某个方向上的单个组织的所有切片求平均，再对所有组织求平均。最终对所有case求平均。
                        case_mean.append(
                            np.mean(np.mean(case_item, axis=0), axis=0))
                self.mean_by_all[direction][metric_name] = {
                    'mean': np.mean(case_mean),
                    'std': np.std(case_mean)
                }

        # flow
        for direction in ['ed_to_es', 'es_to_ed']:
            for metric_name in ['BE', 'Jacobian']:
                case_mean = []
                for dataset_name in self.details:
                    dataset = self.details[dataset_name]
                    case_mean += [
                        np.mean(
                            dataset[case_no]['flow'][direction][metric_name])
                        for case_no in dataset
                    ]
                self.mean_by_all[direction][metric_name] = {
                    'mean': np.mean(case_mean),
                    'std': np.std(case_mean)
                }
    
    def MMs_Vendor_meanByAnotomicalOfDataset(self):

        self.MMs_vendor_mean_by_anotomical = {}

        for dataset_name in self.details:
            dataset = self.details[dataset_name]
            self.MMs_vendor_mean_by_anotomical[dataset_name] = {}
            for anotomical in self.info[dataset_name]:
                self.MMs_vendor_mean_by_anotomical[dataset_name][anotomical] = {}
                for direction in ['undef', 'ed_to_es', 'es_to_ed']:
                    self.MMs_vendor_mean_by_anotomical[dataset_name][anotomical][direction] = {}
                    for metric_name in ['Dice', 'SymAPD', 'HD', 'APDgt_pred', 'APDpred_gt']:
                        A_case_mean = []
                        B_case_mean = []
                        C_case_mean = []
                        D_case_mean = []
                        self.MMs_vendor_mean_by_anotomical[dataset_name][anotomical][direction][metric_name] = {}
                        for case_no in dataset:
                            vendor = self.info_list[self.info_list[:, 0] == case_no, ...][0][2]
                            if vendor == 'A':
                                A_case_mean.append(np.mean(dataset[case_no][anotomical][direction][metric_name]))
                            elif vendor == 'B':
                                B_case_mean.append(np.mean(dataset[case_no][anotomical][direction][metric_name]))
                            elif vendor == 'C':
                                C_case_mean.append(np.mean(dataset[case_no][anotomical][direction][metric_name]))
                            else:
                                D_case_mean.append(np.mean(dataset[case_no][anotomical][direction][metric_name]))

                        self.MMs_vendor_mean_by_anotomical[dataset_name][anotomical][direction][metric_name]['A'] = {
                                'mean': np.mean(A_case_mean),
                                'std': np.std(A_case_mean)
                            }
                        self.MMs_vendor_mean_by_anotomical[dataset_name][anotomical][direction][metric_name]['B'] = {
                                'mean': np.mean(B_case_mean),
                                'std': np.std(B_case_mean)
                            }
                        self.MMs_vendor_mean_by_anotomical[dataset_name][anotomical][direction][metric_name]['C'] = {
                                'mean': np.mean(C_case_mean),
                                'std': np.std(C_case_mean)
                            }
                        self.MMs_vendor_mean_by_anotomical[dataset_name][anotomical][direction][metric_name]['D'] = {
                                'mean': np.mean(D_case_mean),
                                'std': np.std(D_case_mean)
                            }
    
    def MMs_Vendor_meanByDataset(self):
        # first average in a case
        # then average along with cases

        self.MMs_vendor_mean_by_dataset = {}

        for dataset_name in self.details:
            dataset = self.details[dataset_name]
            self.MMs_vendor_mean_by_dataset[dataset_name] = {}
            for direction in ['undef', 'ed_to_es', 'es_to_ed']:
                self.MMs_vendor_mean_by_dataset[dataset_name][direction] = {}
                for metric_name in ['Dice', 'SymAPD', 'HD', 'APDgt_pred', 'APDpred_gt']:
                    A_case_mean = []
                    B_case_mean = []
                    C_case_mean = []
                    D_case_mean = []
                    self.MMs_vendor_mean_by_dataset[dataset_name][direction][metric_name] = {}
                    for case_no in dataset:
                        vendor = self.info_list[self.info_list[:, 0] == case_no, ...][0][2]
                        case = dataset[case_no]
                        case_item = np.array([case[anotomical][direction][metric_name] for anotomical in self.info[dataset_name]])
                        if vendor == 'A':
                            A_case_mean.append(np.mean(np.mean(case_item, axis=0), axis=0))
                        elif vendor == 'B':
                            B_case_mean.append(np.mean(np.mean(case_item, axis=0), axis=0))
                        elif vendor == 'C':
                            C_case_mean.append(np.mean(np.mean(case_item, axis=0), axis=0))
                        else:
                            D_case_mean.append(np.mean(np.mean(case_item, axis=0), axis=0))

                    self.MMs_vendor_mean_by_dataset[dataset_name][direction][metric_name]['A'] = {'mean': np.mean(A_case_mean), 'std': np.std(A_case_mean)}
                    self.MMs_vendor_mean_by_dataset[dataset_name][direction][metric_name]['B'] = {'mean': np.mean(B_case_mean), 'std': np.std(B_case_mean)}
                    self.MMs_vendor_mean_by_dataset[dataset_name][direction][metric_name]['C'] = {'mean': np.mean(C_case_mean), 'std': np.std(C_case_mean)}
                    self.MMs_vendor_mean_by_dataset[dataset_name][direction][metric_name]['D'] = {'mean': np.mean(D_case_mean), 'std': np.std(D_case_mean)}
            
            # flow
            for direction in ['ed_to_es', 'es_to_ed']:
                for metric_name in ['BE', 'Jacobian']:
                    A_case_mean = []
                    B_case_mean = []
                    C_case_mean = []
                    D_case_mean = []
                    self.MMs_vendor_mean_by_dataset[dataset_name][direction][metric_name] = {}
                    for case_no in dataset:
                        vendor = self.info_list[self.info_list[:, 0] == case_no, ...][0][2]
                        if vendor == 'A':
                            A_case_mean.append(np.mean(dataset[case_no]['flow'][direction][metric_name]))
                        elif vendor == 'B':
                            B_case_mean.append(np.mean(dataset[case_no]['flow'][direction][metric_name]))
                        elif vendor == 'C':
                            C_case_mean.append(np.mean(dataset[case_no]['flow'][direction][metric_name]))
                        else:
                            D_case_mean.append(np.mean(dataset[case_no]['flow'][direction][metric_name]))
                    
                    self.MMs_vendor_mean_by_dataset[dataset_name][direction][metric_name]['A'] = {'mean': np.mean(A_case_mean), 'std': np.std(A_case_mean)}
                    self.MMs_vendor_mean_by_dataset[dataset_name][direction][metric_name]['B'] = {'mean': np.mean(B_case_mean), 'std': np.std(B_case_mean)}
                    self.MMs_vendor_mean_by_dataset[dataset_name][direction][metric_name]['C'] = {'mean': np.mean(C_case_mean), 'std': np.std(C_case_mean)}
                    self.MMs_vendor_mean_by_dataset[dataset_name][direction][metric_name]['D'] = {'mean': np.mean(D_case_mean), 'std': np.std(D_case_mean)}

    def MMs_Vendor_MeanByAll(self):
        self.MMs_vendor_mean_by_all = {}

        for direction in ['undef', 'ed_to_es', 'es_to_ed']:
            self.MMs_vendor_mean_by_all[direction] = {}
            for metric_name in ['Dice', 'SymAPD', 'HD', 'APDgt_pred', 'APDpred_gt']:
                self.MMs_vendor_mean_by_all[direction][metric_name] = {}
                for dataset_name in self.details:
                    A_case_mean = []
                    B_case_mean = []
                    C_case_mean = []
                    D_case_mean = []
                    dataset = self.details[dataset_name]
                    for case_no in dataset:
                        vendor = self.info_list[self.info_list[:, 0] == case_no, ...][0][2]
                        case = dataset[case_no]
                        case_item = np.array([case[anotomical][direction][metric_name] for anotomical in self.info[dataset_name]])
                        if vendor == 'A':
                            A_case_mean.append(np.mean(np.mean(case_item, axis=0), axis=0))
                        elif vendor == 'B':
                            B_case_mean.append(np.mean(np.mean(case_item, axis=0), axis=0))
                        elif vendor == 'C':
                            C_case_mean.append(np.mean(np.mean(case_item, axis=0), axis=0))
                        else:
                            D_case_mean.append(np.mean(np.mean(case_item, axis=0), axis=0))
                
                self.MMs_vendor_mean_by_all[direction][metric_name]['A'] = {'mean': np.mean(A_case_mean), 'std': np.std(A_case_mean)}
                self.MMs_vendor_mean_by_all[direction][metric_name]['B'] = {'mean': np.mean(B_case_mean), 'std': np.std(B_case_mean)}
                self.MMs_vendor_mean_by_all[direction][metric_name]['C'] = {'mean': np.mean(C_case_mean), 'std': np.std(C_case_mean)}
                self.MMs_vendor_mean_by_all[direction][metric_name]['D'] = {'mean': np.mean(D_case_mean), 'std': np.std(D_case_mean)}

        for direction in ['ed_to_es', 'es_to_ed']:
            for metric_name in ['BE', 'Jacobian']:
                self.MMs_vendor_mean_by_all[direction][metric_name] = {}
                for dataset_name in self.details:
                    A_case_mean = []
                    B_case_mean = []
                    C_case_mean = []
                    D_case_mean = []
                    dataset = self.details[dataset_name]
                    for case_no in dataset:
                        vendor = self.info_list[self.info_list[:, 0] == case_no, ...][0][2]
                        if vendor == 'A':
                            A_case_mean.append(np.mean(dataset[case_no]['flow'][direction][metric_name]))
                        elif vendor == 'B':
                            B_case_mean.append(np.mean(dataset[case_no]['flow'][direction][metric_name]))
                        elif vendor == 'C':
                            C_case_mean.append(np.mean(dataset[case_no]['flow'][direction][metric_name]))
                        else:
                            D_case_mean.append(np.mean(dataset[case_no]['flow'][direction][metric_name]))

                self.MMs_vendor_mean_by_all[direction][metric_name]['A'] = {'mean': np.mean(A_case_mean), 'std': np.std(A_case_mean)}
                self.MMs_vendor_mean_by_all[direction][metric_name]['B'] = {'mean': np.mean(B_case_mean), 'std': np.std(B_case_mean)}
                self.MMs_vendor_mean_by_all[direction][metric_name]['C'] = {'mean': np.mean(C_case_mean), 'std': np.std(C_case_mean)}
                self.MMs_vendor_mean_by_all[direction][metric_name]['D'] = {'mean': np.mean(D_case_mean), 'std': np.std(D_case_mean)}

    def collectionOfMetricOnTissues(self, model_name):
        self.collection_of_Metric = {}
        for tissues in ['Rv', 'Lv', 'LvBp', 'LvMyo']:
            self.collection_of_Metric[tissues] = {}
            for metric_name in ['Dice', 'SymAPD', 'HD']:
                self.collection_of_Metric[tissues][metric_name] = []
        
        for dataset_name in self.details:
            if self.details[dataset_name] == {}:
                continue
            dataset = self.details[dataset_name]
            for case_no in dataset:
                case = dataset[case_no]
                for tissues in ['Rv', 'Lv', 'LvBp', 'LvMyo']:
                    if case.get(tissues) == None:
                        continue
                    else:
                        for metric_name in ['Dice', 'SymAPD', 'HD']:
                            self.collection_of_Metric[tissues][metric_name].extend(case[tissues]['ed_to_es'][metric_name])
        # 写入文件
        dataset_name = 'MMs' if self.details['M&M'] != {} else 'hybrid'
        for metric_name in ['Dice', 'SymAPD', 'HD']:
            temp_list = []
            for tissues in ['Rv', 'Lv', 'LvBp', 'LvMyo']:
                temp_list.append(self.collection_of_Metric[tissues][metric_name])
            for i, tissues in zip(range(len(temp_list)), ['Rv', 'Lv', 'LvBp', 'LvMyo']):
                temp = np.array(temp_list[i])
                np.savetxt('tissues_metrics/' + model_name + '_' + dataset_name + '_' + tissues + '_' + metric_name + '.txt', temp)
            
        print('1')

    def bar_pic_by_metric(self, model_name, train_mode, test_mode):
        # Hybrid(10%), Hybrid(15%), MMs(5%), MMs(15%), Truth
        # 2个数据集，5种非均匀点的生成策略
        # 画APD，Jacobian和Dice的Bar图

        self.collection_of_Metric = {}
        for metric_name in ['SymAPD', 'Jacobian', 'Dice']:
            self.collection_of_Metric[metric_name] = []

        for dataset_name in self.details:
            if self.details[dataset_name] == {}:
                continue
            dataset = self.details[dataset_name]
            for case_no in dataset:
                case = dataset[case_no]
                for metric_name in ['SymAPD', 'Dice']:
                    case_item = np.array([case[anotomical]['ed_to_es'][metric_name] for anotomical in self.info[dataset_name]])
                    case_item = np.mean(np.mean(case_item, axis=0), axis=0)
                    self.collection_of_Metric[metric_name].append(case_item)
                case_item = np.mean(case['flow']['ed_to_es']['Jacobian'])
                self.collection_of_Metric['Jacobian'].append(case_item)
        
         # 写入文件
        dataset_name = 'MMs' if self.details['M&M'] != {} else 'Hybrid'
        data = []
        for metric_name in ['SymAPD', 'Jacobian', 'Dice']:
            item = np.array(self.collection_of_Metric[metric_name])
            data.append(item)
        data = np.stack(data).T
        np.savetxt('{}_{}{}_{}_SymAPD_Jacobian_Dice'.format(model_name, dataset_name, train_mode, test_mode), data)
        print('1')

    
    def mean(self):
        self.meanByAnotomicalOfDataset()
        self.meanByDataset()
        self.meanByAll()
        if self.details['M&M'] != {}:
            self.info_list = pd.read_csv('../shsdata/dataset2D_MnMs/ZiYu_MMs_info.txt', sep=' ', header=None).values
            self.MMs_Vendor_meanByAnotomicalOfDataset()
            self.MMs_Vendor_meanByDataset()
            self.MMs_Vendor_MeanByAll()
        return self.mean_by_all['ed_to_es']['Dice']

    def cellValue(self, ws, row, column, value):
        alignment = Alignment(vertical='center', horizontal='center')
        cell = ws.cell(row=row, column=column)
        cell.alignment = alignment
        cell.value = value

    def getWidth(self, name: str):
        width = 0
        for c in name:
            if c.isdigit():
                width += 1
            elif c.isupper():
                width += 1.2
            else:
                width += 1.1

        return width

    def autoSetWidth(self, ws):
        col_width = []
        #获取每一列的内容的最大宽度
        i = 0
        # 每列
        for col in ws.columns:
            # 每行
            for j in range(len(col)):
                if j in [0, 1, 2, 9]:
                    continue
                if j == 3:
                    # 数组增加一个元素
                    col_width.append(self.getWidth(str(col[j].value)))
                else:
                    # 获得每列中的内容的最大宽度
                    if col_width[i] < self.getWidth(str(col[j].value)):
                        col_width[i] = self.getWidth(str(col[j].value))
            i = i + 1

        #设置列宽
        for i in range(len(col_width)):
            # 根据列的数字返回字母
            col_letter = get_column_letter(i + 1)
            # 当宽度大于100，宽度设置为100
            if col_width[i] > 100:
                ws.column_dimensions[col_letter].width = 100
            # 只有当宽度大于10，才设置列宽
            elif col_width[i] > 10:
                ws.column_dimensions[col_letter].width = col_width[i] + 2

    def initWorksheet(self, ws: worksheet.worksheet.Worksheet):
        alignment = Alignment(vertical='center', horizontal='center')
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)
        self.cellValue(ws, 1, 1, 'Network')
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=6)
        self.cellValue(ws, 2, 1, 'Name')
        metric_name = ['BE', 'Jacobian', 'Dice', 'HD', 'SymAPD', 'APDgt_pred', 'APDpred_gt']
        dataset_name = ['ACDC', 'York', 'MICCAI', 'ALL', 'M&M', 'M&M_A', 'M&M_B', 'M&M_C', 'M&M_D']

        ws.merge_cells(start_row=4, start_column=1, end_row=5, end_column=1)
        self.cellValue(ws, 4, 1, dataset_name[0])
        self.cellValue(ws, 6, 1, dataset_name[1])
        self.cellValue(ws, 7, 1, dataset_name[2])
        self.cellValue(ws, 8, 1, dataset_name[3])
        self.cellValue(ws, 9, 1, dataset_name[4])

        for i in range(7):
            self.cellValue(ws, 4, 2 + i, metric_name[i])

        ws.merge_cells(start_row=11, start_column=1, end_row=15, end_column=1)
        self.cellValue(ws, 11, 1, dataset_name[0])
        ws.merge_cells(start_row=16, start_column=1, end_row=18, end_column=1)
        self.cellValue(ws, 16, 1, dataset_name[1])
        ws.merge_cells(start_row=19, start_column=1, end_row=22, end_column=1)
        self.cellValue(ws, 19, 1, dataset_name[4])

        for i in range(5):
            self.cellValue(ws, 11, 3 + i, metric_name[i + 2])

        anatomical_list = list(self.info['ACDC'])
        t = 12
        for anatomical_name in anatomical_list:
            self.cellValue(ws, t, 2, anatomical_name)
            t = t + 1
        anatomical_list = list(self.info['York'])
        for anatomical_name in anatomical_list:
            self.cellValue(ws, t, 2, anatomical_name)
            t = t + 1
        anatomical_list = list(self.info['M&M'])
        for anatomical_name in anatomical_list:
            self.cellValue(ws, t, 2, anatomical_name)
            t = t + 1

        # 添加M&Ms数据集各厂商的测试结果
        ws.merge_cells(start_row=24, start_column=1, end_row=25, end_column=1)
        self.cellValue(ws, 24, 1, dataset_name[5])
        self.cellValue(ws, 26, 1, dataset_name[6])
        self.cellValue(ws, 27, 1, dataset_name[7])
        self.cellValue(ws, 28, 1, dataset_name[8])

        for i in range(7):
            self.cellValue(ws, 24, 2 + i, metric_name[i])

        ws.merge_cells(start_row=30, start_column=1, end_row=34, end_column=1)
        self.cellValue(ws, 30, 1, dataset_name[5])
        ws.merge_cells(start_row=35, start_column=1, end_row=38, end_column=1)
        self.cellValue(ws, 35, 1, dataset_name[6])
        ws.merge_cells(start_row=39, start_column=1, end_row=42, end_column=1)
        self.cellValue(ws, 39, 1, dataset_name[7])
        ws.merge_cells(start_row=43, start_column=1, end_row=46, end_column=1)
        self.cellValue(ws, 43, 1, dataset_name[8])

        for i in range(5):
            self.cellValue(ws, 30, 3 + i, metric_name[i + 2])

        t = 31
        for anatomical_name in anatomical_list:
            self.cellValue(ws, t, 2, anatomical_name)
            self.cellValue(ws, t + 4, 2, anatomical_name)
            self.cellValue(ws, t + 8, 2, anatomical_name)
            self.cellValue(ws, t + 12, 2, anatomical_name)
            t = t + 1
        #

    def setWorksheet(self, ws: worksheet.worksheet.Worksheet,
                     network, name, direction):
        alignment = Alignment(vertical='center', horizontal='center')
        self.cellValue(ws, 2, 2, name)
        self.cellValue(ws, 1, 2, network)

        metric_name = ['BE', 'Jacobian', 'Dice', 'HD', 'SymAPD', 'APDgt_pred', 'APDpred_gt']
        metric_format = ['%.2f(%.2f)', '%.2f(%.2f)', '%.4f(%.2f)', '%.2f(%.2f)','%.2f(%.2f)', '%.2f(%.2f)', '%.2f(%.2f)']
        dataset_name = ['ACDC', 'York', 'MICCAI', 'ALL', 'M&M']
        for i in range(7):
            if metric_name[i] in self.mean_by_dataset[dataset_name[0]][direction]:
                metric_value = self.mean_by_dataset[dataset_name[0]][direction][metric_name[i]]
                value = metric_format[i] % (metric_value['mean'], metric_value['std'])
                self.cellValue(ws, 5, i + 2, value)
        for i in range(7):
            if metric_name[i] in self.mean_by_dataset[dataset_name[1]][direction]:
                metric_value = self.mean_by_dataset[dataset_name[1]][direction][metric_name[i]]
                value = metric_format[i] % (metric_value['mean'], metric_value['std'])
                self.cellValue(ws, 6, i + 2, value)
        for i in range(7):
            if metric_name[i] in self.mean_by_dataset[dataset_name[2]][direction]:
                metric_value = self.mean_by_dataset[dataset_name[2]][direction][metric_name[i]]
                value = metric_format[i] % (metric_value['mean'], metric_value['std'])
                self.cellValue(ws, 7, i + 2, value)
        for i in range(3, 10):
            if metric_name[i - 3] in self.mean_by_all[direction]:
                metric_value = self.mean_by_all[direction][metric_name[i - 3]]
                value = metric_format[i - 3] % (metric_value['mean'], metric_value['std'])
                self.cellValue(ws, 8, i - 1, value)
        for i in range(7):
            if metric_name[i] in self.mean_by_dataset[dataset_name[4]][direction]:
                metric_value = self.mean_by_dataset[dataset_name[4]][direction][metric_name[i]]
                value = metric_format[i] % (metric_value['mean'], metric_value['std'])
                self.cellValue(ws, 9, i + 2, value)

        metric_name = ['Dice', 'HD', 'SymAPD', 'APDgt_pred', 'APDpred_gt']
        metric_format = ['%.4f(%.2f)', '%.2f(%.2f)', '%.2f(%.2f)', '%.2f(%.2f)', '%.2f(%.2f)']


        t = 12
        anatomical_list = list(self.info[dataset_name[0]])
        for anatomical_name in anatomical_list:
            for i in range(5):
                metric_value = self.mean_by_anotomical[dataset_name[0]][anatomical_name][direction][metric_name[i]]
                value = metric_format[i] % (metric_value['mean'], metric_value['std'])
                self.cellValue(ws, t, i + 3, value)
            t = t + 1

        anatomical_list = list(self.info[dataset_name[1]])
        for anatomical_name in anatomical_list:
            for i in range(5):
                metric_value = self.mean_by_anotomical[dataset_name[1]][anatomical_name][direction][metric_name[i]]
                value = metric_format[i] % (metric_value['mean'], metric_value['std'])
                self.cellValue(ws, t, i + 3, value)
            t = t + 1

        anatomical_list = list(self.info[dataset_name[4]])
        for anatomical_name in anatomical_list:
            for i in range(5):
                metric_value = self.mean_by_anotomical[dataset_name[4]][anatomical_name][direction][metric_name[i]]
                value = metric_format[i] % (metric_value['mean'], metric_value['std'])
                self.cellValue(ws, t, i + 3, value)
            t = t + 1

        # 添加M&Ms数据集各厂商的测试结果
        if self.details['M&M'] != {}:
            metric_name = ['BE', 'Jacobian', 'Dice', 'HD', 'SymAPD', 'APDgt_pred', 'APDpred_gt']
            metric_format = ['%.2f(%.2f)', '%.2f(%.2f)', '%.4f(%.2f)', '%.2f(%.2f)','%.2f(%.2f)', '%.2f(%.2f)', '%.2f(%.2f)']

            for i in range(7):
                if metric_name[i] in self.MMs_vendor_mean_by_dataset[dataset_name[4]][direction]:
                    metric_value = self.MMs_vendor_mean_by_dataset[dataset_name[4]][direction][metric_name[i]]['A']
                    value = metric_format[i] % (metric_value['mean'], metric_value['std'])
                    self.cellValue(ws, 25, i + 2, value)
            
            for i in range(7):
                if metric_name[i] in self.MMs_vendor_mean_by_dataset[dataset_name[4]][direction]:
                    metric_value = self.MMs_vendor_mean_by_dataset[dataset_name[4]][direction][metric_name[i]]['B']
                    value = metric_format[i] % (metric_value['mean'], metric_value['std'])
                    self.cellValue(ws, 26, i + 2, value)
            
            for i in range(7):
                if metric_name[i] in self.MMs_vendor_mean_by_dataset[dataset_name[4]][direction]:
                    metric_value = self.MMs_vendor_mean_by_dataset[dataset_name[4]][direction][metric_name[i]]['C']
                    value = metric_format[i] % (metric_value['mean'], metric_value['std'])
                    self.cellValue(ws, 27, i + 2, value)
            
            for i in range(7):
                if metric_name[i] in self.MMs_vendor_mean_by_dataset[dataset_name[4]][direction]:
                    metric_value = self.MMs_vendor_mean_by_dataset[dataset_name[4]][direction][metric_name[i]]['D']
                    value = metric_format[i] % (metric_value['mean'], metric_value['std'])
                    self.cellValue(ws, 28, i + 2, value)

            metric_name = ['Dice', 'HD', 'SymAPD', 'APDgt_pred', 'APDpred_gt']
            metric_format = ['%.4f(%.2f)', '%.2f(%.2f)', '%.2f(%.2f)', '%.2f(%.2f)', '%.2f(%.2f)']

            t = 31
            for anatomical_name in anatomical_list:
                for i in range(5):
                    metric_value = self.MMs_vendor_mean_by_anotomical[dataset_name[4]][anatomical_name][direction][metric_name[i]]['A']
                    value = metric_format[i] % (metric_value['mean'], metric_value['std'])
                    self.cellValue(ws, t, i + 3, value)
                    metric_value = self.MMs_vendor_mean_by_anotomical[dataset_name[4]][anatomical_name][direction][metric_name[i]]['B']
                    value = metric_format[i] % (metric_value['mean'], metric_value['std'])
                    self.cellValue(ws, t + 4, i + 3, value)
                    metric_value = self.MMs_vendor_mean_by_anotomical[dataset_name[4]][anatomical_name][direction][metric_name[i]]['C']
                    value = metric_format[i] % (metric_value['mean'], metric_value['std'])
                    self.cellValue(ws, t + 8, i + 3, value)
                    metric_value = self.MMs_vendor_mean_by_anotomical[dataset_name[4]][anatomical_name][direction][metric_name[i]]['D']
                    value = metric_format[i] % (metric_value['mean'], metric_value['std'])
                    self.cellValue(ws, t + 12, i + 3, value)
                t = t + 1
        #

    def saveAsExcel(self, network, name, excel_save_path, logger=None):
        wb = Workbook()
        wb_name = '%s.xlsx' % (name)
        wb_save_path = os.path.join(excel_save_path, wb_name)

        if logger != None:
            ws0 = wb.create_sheet(title='training')
            alignment = Alignment(vertical='center', horizontal='center')
            res = []
            with open(os.path.join(logger.log_dir, 'log.csv'), 'r', encoding='utf-8', newline='') as fp:
                reader = csv.reader(fp)
                next(reader)
                for row in reader:
                    row = [float(x) for x in row]
                    res.append(row)
            res = np.mean(res, axis=0)
            for i in range(res.size):
                self.cellValue(ws0, 1, i + 1, res[i])
            
            last_row = None
            with open(os.path.join(logger.log_dir, 'update_info.csv'), 'r', encoding='utf-8', newline='') as fp:
                reader = csv.reader(fp)
                for row in reader:
                    last_row = row
            for i in range(len(last_row)):
                self.cellValue(ws0, 2, i + 1, last_row[i])

        ws1 = wb.active
        ws1.title = 'ed_to_es'
        self.initWorksheet(ws1)
        self.setWorksheet(ws1, network, name, 'ed_to_es')
        self.autoSetWidth(ws1)

        ws2 = wb.create_sheet(title='es_to_ed')
        self.initWorksheet(ws2)
        self.setWorksheet(ws2, network, name, 'es_to_ed')
        self.autoSetWidth(ws2)

        ws3 = wb.create_sheet(title='undef')
        self.initWorksheet(ws3)
        self.setWorksheet(ws3, 'undef', '', 'undef')
        self.autoSetWidth(ws3)

        wb.save(wb_save_path)

    def output(self):
        # pprint.pprint(self.mean_by_anotomical)
        # pprint.pprint(self.mean_by_dataset)
        pprint.pprint(self.mean_by_all)
        print('Done')